"""
export.py — Exportação de dados do Auroque para Excel e CSV
Cobertura total: animais, pesagens, custos, DRE, veterinário, etc.
"""
import io
import logging
from datetime import datetime

_log = logging.getLogger("auroque.export")

try:
    import openpyxl
    from openpyxl.styles import (Font, PatternFill, Alignment,
                                  Border, Side)
    from openpyxl.utils import get_column_letter
    _EXCEL_OK = True
except ImportError:
    _EXCEL_OK = False
    _log.warning("openpyxl nao instalado — export Excel indisponivel")


# ── HELPERS ───────────────────────────────────────────────────

def _estilizar_header(ws, row=1, cor_fundo="1B4332", cor_texto="F5F0E8"):
    """Aplica estilo de cabeçalho verde Auroque."""
    if not _EXCEL_OK:
        return
    fill  = PatternFill("solid", fgColor=cor_fundo)
    fonte = Font(bold=True, color=cor_texto)
    alin  = Alignment(horizontal="center", vertical="center")
    for cell in ws[row]:
        cell.fill      = fill
        cell.font      = fonte
        cell.alignment = alin


def _auto_width(ws):
    """Ajusta largura das colunas automaticamente."""
    if not _EXCEL_OK:
        return
    for col in ws.columns:
        max_w = 0
        col_l = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_w = max(max_w, len(str(cell.value or "")))
            except Exception:
                pass
        ws.column_dimensions[col_l].width = min(max(max_w + 2, 10), 50)


def _df_para_sheet(wb, df, nome_aba, cor="1B4332"):
    """Adiciona DataFrame como aba no workbook."""
    if df is None or (hasattr(df, "empty") and df.empty):
        return
    ws = wb.create_sheet(title=nome_aba[:31])
    # Cabeçalho
    for col_idx, col_name in enumerate(df.columns, 1):
        ws.cell(row=1, column=col_idx, value=str(col_name))
    # Dados
    for row_idx, row in enumerate(df.itertuples(index=False), 2):
        for col_idx, value in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    _estilizar_header(ws, cor_fundo=cor)
    _auto_width(ws)
    return ws


# ── EXPORTAÇÕES ESPECÍFICAS ───────────────────────────────────

def exportar_animais(owner_id) -> bytes:
    """Exporta lista de animais para Excel."""
    import pandas as pd
    from database import listar_lotes, listar_animais_por_lote

    if not _EXCEL_OK:
        return _exportar_csv_fallback_animais(owner_id)

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    lotes = listar_lotes(owner_id=owner_id) or []
    for lote in lotes:
        lid, nome_lote = lote[0], lote[1]
        animais = listar_animais_por_lote(lid) or []
        if not animais:
            continue
        rows = []
        for a in animais:
            rows.append({
                "Identificação": a[1] if len(a)>1 else "",
                "Raça":          a[2] if len(a)>2 else "",
                "Sexo":          a[3] if len(a)>3 else "",
                "Idade (meses)": a[4] if len(a)>4 else "",
                "Peso entrada":  a[5] if len(a)>5 else "",
                "Status":        a[7] if len(a)>7 else "ATIVO",
            })
        df = pd.DataFrame(rows)
        _df_para_sheet(wb, df, nome_lote[:31])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def exportar_pesagens(owner_id) -> bytes:
    """Exporta pesagens de todos os animais."""
    import pandas as pd
    from database import listar_lotes, listar_animais_por_lote, listar_pesagens

    wb = openpyxl.Workbook() if _EXCEL_OK else None
    rows_all = []

    lotes = listar_lotes(owner_id=owner_id) or []
    for lote in lotes:
        lid, nome_lote = lote[0], lote[1]
        for animal in (listar_animais_por_lote(lid) or []):
            aid, ident = animal[0], animal[1]
            for p in (listar_pesagens(aid) or []):
                rows_all.append({
                    "Lote":           nome_lote,
                    "Identificação":  ident,
                    "Peso (kg)":      p[2] if len(p)>2 else "",
                    "Data":           str(p[3])[:10] if len(p)>3 else "",
                })

    df = pd.DataFrame(rows_all)
    if _EXCEL_OK and wb:
        _df_para_sheet(wb, df, "Pesagens")
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
    else:
        return df.to_csv(index=False).encode("utf-8")


def exportar_financeiro(owner_id) -> bytes:
    """Exporta DRE, custos e vendas em abas separadas."""
    import pandas as pd
    from database import (listar_custos_lote, listar_vendas_lote,
                          listar_lotes)

    if not _EXCEL_OK:
        _log.warning("openpyxl nao disponivel")
        return b""

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    lotes = listar_lotes(owner_id=owner_id) or []

    # Aba custos
    rows_custos = []
    for lote in lotes:
        lid, nome = lote[0], lote[1]
        for c in (listar_custos_lote(lid) or []):
            rows_custos.append({
                "Lote":       nome,
                "Categoria":  c[1] if len(c)>1 else "",
                "Descrição":  c[2] if len(c)>2 else "",
                "Valor":      c[3] if len(c)>3 else 0,
                "Data":       str(c[4])[:10] if len(c)>4 else "",
            })
    _df_para_sheet(wb, pd.DataFrame(rows_custos), "Custos")

    # Aba vendas
    rows_vendas = []
    for lote in lotes:
        lid, nome = lote[0], lote[1]
        for v in (listar_vendas_lote(lid) or []):
            rows_vendas.append({
                "Lote":        nome,
                "Data venda":  str(v[1])[:10] if len(v)>1 else "",
                "Preço/kg":    v[2] if len(v)>2 else 0,
                "Peso total":  v[3] if len(v)>3 else 0,
                "Frigorífico": v[4] if len(v)>4 else "",
            })
    _df_para_sheet(wb, pd.DataFrame(rows_vendas), "Vendas")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def exportar_veterinario(owner_id) -> bytes:
    """Exporta receituário, vacinas e ocorrências."""
    import pandas as pd
    from database import listar_receitas, listar_vacinas_agenda, listar_ocorrencias

    if not _EXCEL_OK:
        return b""

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Receitas
    receitas = listar_receitas(fazenda_owner_id=owner_id) or []
    rows_r = []
    for r in receitas:
        rows_r.append({
            "Medicamento": r[3] if len(r)>3 else "",
            "Dose":        r[4] if len(r)>4 else "",
            "Via":         r[5] if len(r)>5 else "",
            "Data":        str(r[6])[:10] if len(r)>6 else "",
            "Carência":    r[7] if len(r)>7 else "",
        })
    _df_para_sheet(wb, pd.DataFrame(rows_r), "Receituário")

    # Vacinas
    vacinas = listar_vacinas_agenda(owner_id=owner_id) or []
    rows_v = []
    for v in vacinas:
        rows_v.append({
            "Vacina":         v[1] if len(v)>1 else "",
            "Data prevista":  str(v[2])[:10] if len(v)>2 else "",
            "Data realizada": str(v[3])[:10] if len(v)>3 else "",
            "Status":         v[4] if len(v)>4 else "",
        })
    _df_para_sheet(wb, pd.DataFrame(rows_v), "Vacinas", cor="40916C")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def exportar_tudo(owner_id) -> bytes:
    """Exporta todos os dados em um único arquivo Excel."""
    import pandas as pd
    from database import (listar_lotes, listar_animais_por_lote,
                          listar_pesagens, listar_custos_lote,
                          listar_vendas_lote, listar_ocorrencias,
                          listar_receitas, listar_vacinas_agenda)

    if not _EXCEL_OK:
        return b""

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    lotes = listar_lotes(owner_id=owner_id) or []

    # Aba: Animais
    rows = []
    for lote in lotes:
        lid, nome = lote[0], lote[1]
        for a in (listar_animais_por_lote(lid) or []):
            rows.append({"Lote": nome, "ID": a[1] if len(a)>1 else "",
                         "Raça": a[2] if len(a)>2 else "",
                         "Sexo": a[3] if len(a)>3 else "",
                         "Peso entrada": a[5] if len(a)>5 else "",
                         "Status": a[7] if len(a)>7 else "ATIVO"})
    _df_para_sheet(wb, pd.DataFrame(rows), "Animais")

    # Aba: Pesagens
    rows = []
    for lote in lotes:
        lid, nome = lote[0], lote[1]
        for a in (listar_animais_por_lote(lid) or []):
            for p in (listar_pesagens(a[0]) or []):
                rows.append({"Lote": nome, "Animal": a[1] if len(a)>1 else "",
                             "Peso": p[2] if len(p)>2 else "",
                             "Data": str(p[3])[:10] if len(p)>3 else ""})
    _df_para_sheet(wb, pd.DataFrame(rows), "Pesagens", cor="40916C")

    # Aba: Custos
    rows = []
    for lote in lotes:
        lid, nome = lote[0], lote[1]
        for c in (listar_custos_lote(lid) or []):
            rows.append({"Lote": nome,
                         "Categoria": c[1] if len(c)>1 else "",
                         "Descrição": c[2] if len(c)>2 else "",
                         "Valor": c[3] if len(c)>3 else 0,
                         "Data": str(c[4])[:10] if len(c)>4 else ""})
    _df_para_sheet(wb, pd.DataFrame(rows), "Custos")

    # Aba: Vendas
    rows = []
    for lote in lotes:
        lid, nome = lote[0], lote[1]
        for v in (listar_vendas_lote(lid) or []):
            rows.append({"Lote": nome,
                         "Data": str(v[1])[:10] if len(v)>1 else "",
                         "Preço/kg": v[2] if len(v)>2 else 0,
                         "Peso total": v[3] if len(v)>3 else 0})
    _df_para_sheet(wb, pd.DataFrame(rows), "Vendas")

    # Aba: Ocorrências
    rows = []
    for lote in lotes:
        lid, nome = lote[0], lote[1]
        for a in (listar_animais_por_lote(lid) or []):
            for o in (listar_ocorrencias(a[0]) or []):
                rows.append({"Lote": nome, "Animal": a[1] if len(a)>1 else "",
                             "Tipo": o[2] if len(o)>2 else "",
                             "Data": str(o[1])[:10] if len(o)>1 else "",
                             "Descrição": o[3] if len(o)>3 else ""})
    _df_para_sheet(wb, pd.DataFrame(rows), "Ocorrências")

    buf = io.BytesIO()
    wb.save(buf)
    _log.info("exportar_tudo: %d abas, owner=%s", len(wb.sheetnames), owner_id)
    return buf.getvalue()


def _exportar_csv_fallback_animais(owner_id) -> bytes:
    """Fallback CSV quando openpyxl não está disponível."""
    import pandas as pd
    from database import listar_lotes, listar_animais_por_lote
    rows = []
    for lote in (listar_lotes(owner_id=owner_id) or []):
        for a in (listar_animais_por_lote(lote[0]) or []):
            rows.append({"lote": lote[1], "identificacao": a[1] if len(a)>1 else "",
                         "raca": a[2] if len(a)>2 else ""})
    return pd.DataFrame(rows).to_csv(index=False).encode("utf-8")
